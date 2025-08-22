from django.shortcuts import render, redirect
from django.utils.timezone import now
import os
import csv
from openpyxl import Workbook, load_workbook

LOGS_DIR = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(LOGS_DIR, exist_ok=True)
CSV_PATH = os.path.join(LOGS_DIR, "phish_results.csv")
XLSX_PATH = os.path.join(LOGS_DIR, "phish_results.xlsx")

def login_page(request):
    return render(request, "tracker/login.html")

def submit_login(request):
    if request.method == "POST":
        email = request.POST.get("email", "UNKNOWN")
        ip = request.META.get("REMOTE_ADDR", "UNKNOWN")
        ua = request.META.get("HTTP_USER_AGENT", "UNKNOWN")
        timestamp = now().isoformat()

        # --- CSV Logging ---
        with open(CSV_PATH, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([email, ip, ua, timestamp])

        # --- Excel Logging ---
        if not os.path.exists(XLSX_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["Email", "IP Address", "User-Agent", "Timestamp"])
            wb.save(XLSX_PATH)

        wb = load_workbook(XLSX_PATH)
        ws = wb.active
        ws.append([email, ip, ua, timestamp])
        wb.save(XLSX_PATH)

        return redirect("/training")
    return redirect("/")

def training_page(request):
    return render(request, "tracker/training.html")
